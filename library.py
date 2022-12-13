import openpyxl
from openpyxl import load_workbook
import goto
#-------------------------------------------------------------------------------
#loading and activating workbook
wb = load_workbook('library.xlsx')
ws = wb.active
wb.save('library.xlsx')
#-------------------------------------------------------------------------------
class book:
    def __init__(self, srno , title, ISBN , author , code ):
        self.title = title
        self.author = author
        self.isbn = ISBN
        self.srno = srno
        self.status = "Available"
        self.code = code

    
    def borrow(self):
        self.status = "Borrowed"
        print(str(self.title) + ' Is ' + str(self.status) )

    
    def returnn(self):
        self.status = "Available"
        print(str(self.title) + ' Is ' + str(self.status) )
        
    
    def reserve(self):
        self.status = "Reserved"
        print(str(self.title) + ' Is ' + str(self.status) )

    @classmethod
    def library(cls):
        pass
        
        
class shelf:
    def __init__(self, books_arr   ):
        self.books_arr = books_arr
        
        
    
    def bcount(self):
        self.bcount = str(len(self.books_arr))

    
    def catalog(self):
        i = 1
        print("Sr.No.  -   Title  -  code ")
        for book in self.books_arr:
            print(str(i) + '   |    ' + book + '   |    ' + "book" + str(i) )
            i += 1

            
    
    def add_books(self , book):
        if book not in books_arr:
            self.books_arr.append(book)
            print(book + "Added To Shelf")
        else:
            print("Book Already Exists")
            
            
    
    def remove_books(self, bok):
        if bok in books_arr:
            self.books_arr.remove(bok)
            print(bok + "Removed From Shelf")
        else:
            print("No Such Book")

    #def populate_book(self):
        #populate(args)
#-------------------------------------------------------------------------------
total_rows = ws.max_row
total_column = ws.max_column 
var_for_book = []
row_tuple = []
books_array = []
for j in range(1 , total_rows ):
    row_tuple.append("Book" + str(j))
    var_for_book.append("book" + str(j))
    
    globals()["row_tuples" + str(j)] = []
    
for j in range(1 , total_rows ):
    for i in range(1 , total_column + 1):
        row_obj = ws.cell(row = j +1 , column = i)
            
        globals()["row_tuples" + str(j) ].append(row_obj.value)

for j in range (1 , total_rows ):
    srno_ = globals()["row_tuples" + str(j)][0]
    title_ = globals()["row_tuples" + str(j)][1]
    isbn_ = globals()["row_tuples" + str(j)][2]
    author_ = globals()["row_tuples" + str(j)][3]
    code_ = globals()["row_tuples" + str(j)][5]
    globals()[var_for_book[j-1]] = book(srno_ , title_ , isbn_ , author_ , code_)
    books_array.append(globals()["row_tuples" + str(j)][1])

book_arr = shelf(books_array) 
#-------------------------------------------------------------------------------

cont = 'y' 
cont_2 = 'y'
cont_3 = 'y'

while cont == 'y':
    user_type = input("Who Are You? Enter l for librarian: u for user \n")
    
    if user_type == 'l':
        while cont_2 == 'y':
            print("COMMANDS: ")
            print(" c For Book Count | v For Viewing Catalog | a To Add Books")
            print(" r To Remove Books | e To Exit ")

            cmd = input("Enter Your Command: \n")
            if cmd == 'c':
                book_count = book_arr.bcount()
                print("There are " + str(book_count) + " Books In This Shelf.")

                
            elif cmd == 'v':
                book_arr.catalog()

            elif cmd == 'a':
                book_to_be_added = input("Enter The Book Title : \n")
                book_arr.add_books(book_to_be_added)

            elif cmd == 'r':
                book_to_be_removed = input("Enter The Book Title: \n")
                book_arr.remove_books(book_to_be_removed)

            elif cmd == 'e':
                cont_2 = 'n'

            else:
                print("Enter Correct Command.")


    elif user_type == 'u':
        print("COMMANDS: ")
        print(" v For Viewing Catalog | b To Borrow Book | r To Reserve Book")
        print(" t To Return Book | e To Exit ")


        while cont_3 == 'y':
        
            cmd2 = input("Enter Your Command: \n")
            if cmd2 == 'b':
                book_arr.catalog()
                
                book_to_borrow = input("Enter the title code: \n")
                
                locals()[book_to_borrow].borrow()
                
            elif cmd2 == 'v':
                book_arr.catalog()

            elif cmd2 == 'r':
                book_arr.catalog()
                book_to_reserve = input("Enter the title code: \n")
                locals()[book_to_reserve].reserve()        

            elif cmd2 == 't':
                book_arr.catalog()
                book_to_return = input("Enter the title code: \n")
                locals()[book_to_return].returnn()        

            elif cmd2 == 'e':
                cont_3 = 'n'
                

            else:
                print("Enter Correct Command")
                


    else:
        print("Enter correct user type")
    print("Thanks for using.")
    cont = input("Do you want to change user? (y/n) \n")        

